#!/usr/bin/env python3
"""Assemble a dated, self-contained "apply from here" bundle for one day's hunt.

A day's work is otherwise scattered: the CVs live under applications/<category>/<folder>/,
the tracker holds the metadata, and the briefing is written by hand. This gathers all of it
into one dated folder you can work straight out of, with nothing else open.

Builds:
  <tasks_dir>/<YYYY-MM-DD>/
      CV_<Company>_<Role>.docx            one per role that has a rendered CV
      CoverLetter_<Company>_<Role>.docx   one per role that has a rendered letter
      <YYYY-MM-DD>-roles.xlsx             that day's rows: status, pay, closing date,
                                          fit score, apply link, and the folder path
      FINDINGS.md                         created as a scaffold if absent; NEVER overwritten,
                                          because the agent's written briefing lives there
      APPLY-TODAY.md                      what to send, and where; line 3 is ALWAYS the
                                          profile-check provenance (`Profile check: N docs,
                                          F failures, <profile> @ <mtime>`), so a bundle can
                                          never look the same whether the gate passed or
                                          never ran
      _work/                              the pipeline's working files, tidied out of the way

Selects rows by tracker `logged_date`, so a role sourced today lands in today's bundle even
if its folder is named differently. Re-running is safe: files are refreshed in place, and an
existing FINDINGS.md is left exactly as it is.

Usage:
  python daily_bundle.py --root <apps_dir> [--date YYYY-MM-DD] [--tasks-dir <dir>]
                         [--status Drafted,Skipped] [--quiet]

Defaults: --date today; --tasks-dir <workspace>/tasks/daily (derived from --root).
Prints the bundle path on the last line so a caller can chain off it.
"""
import argparse
import csv
import datetime
import io
import os
import re
import shutil
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from _lib import preflight, safe_cell, enable_utf8_io  # noqa: E402
import validate_profile as vp  # noqa: E402
enable_utf8_io()

preflight()
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import PatternFill, Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

SHEET_COLUMNS = ["status", "category", "company", "role", "location", "pay",
                 "closing_date", "fit_score", "level_used", "source", "link", "folder_path"]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
URGENT_FILL = PatternFill("solid", fgColor="F4CCCC")   # closes within 7 days
SOON_FILL = PatternFill("solid", fgColor="FCE4D6")     # closes within 14 days


def slug(text, maxlen=44):
    text = (text or "").strip()
    text = re.sub(r"\(.*?\)", "", text)                # drop parenthetical asides
    text = re.sub(r"[^A-Za-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:maxlen] or "unknown"


def parse_date(s):
    try:
        return datetime.date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def read_rows(apps, date, statuses):
    path = os.path.join(apps, "tracker.csv")
    if not os.path.exists(path):
        sys.exit(f"No tracker.csv in {apps} — nothing to bundle.")
    with open(path, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    out = []
    for r in rows:
        if str(r.get("logged_date", ""))[:10] != date:
            continue
        if statuses and str(r.get("status", "")) not in statuses:
            continue
        out.append(r)
    return out


def copy_docs(rows, dest, quiet):
    """Copy each role's rendered CV and cover letter in, under a self-describing name."""
    copied = []
    for r in rows:
        folder = str(r.get("folder_path", ""))
        if not folder or not os.path.isdir(folder):
            continue
        stem = f"{slug(r.get('company'))}_{slug(r.get('role'))}"
        for src_name, prefix in (("CV.docx", "CV"), ("CoverLetter.docx", "CoverLetter")):
            src = os.path.join(folder, src_name)
            if not os.path.exists(src):
                continue
            dst = os.path.join(dest, f"{prefix}_{stem}.docx")
            shutil.copy2(src, dst)
            copied.append(os.path.basename(dst))
            if not quiet:
                print(f"  doc     {os.path.basename(dst)}")
    return copied


# The pipeline's working files. Moved to _work/, never deleted: each is an input to a later
# stage (verify_run reads queries.csv; company_rows.json is what harvest_companies writes).
WORK_FILES = ("landscape.csv", "candidates.csv", "shortlist.csv", "ranked.csv",
              "queries.csv", "to-tailor.csv", "company_rows.json")


def tidy(dest, quiet):
    """Move the pipeline's working files into _work/ so the day folder shows only the work."""
    work = os.path.join(dest, "_work")
    moved = 0
    for name in WORK_FILES + ("jds",):
        src = os.path.join(dest, name)
        if not os.path.exists(src):
            continue
        os.makedirs(work, exist_ok=True)
        target = os.path.join(work, name)
        if os.path.isdir(target):
            shutil.rmtree(target, ignore_errors=True)
        shutil.move(src, target)
        moved += 1
    if moved and not quiet:
        print(f"  tidied  {moved} working file(s) into _work/")
    return moved


def profile_check(rows, apps, quiet):
    """Validate every bundled document against the profile. -> ({folder: [failures]}, note).

    The last mechanical gate before a day is called ready: a document that contradicts the
    profile must not be listed as sendable. Best-effort — no profile, no check, and the
    bundle proceeds; but BROKEN RULES are surfaced in the note, never degraded to "skipped".
    """
    prof = vp.find_profile(os.path.dirname(os.path.abspath(apps)))
    if prof is None:
        if not quiet:
            print("  profile check SKIPPED — no profile found under profiles/")
        return {}, "Profile check: SKIPPED — no profile under profiles/"
    stamp = datetime.datetime.fromtimestamp(os.path.getmtime(prof)).strftime("%Y-%m-%d %H:%M")
    ptext = io.open(prof, encoding="utf-8").read()
    try:
        rules = vp.parse_rules(ptext, vp.lanes_for(prof))
    except ValueError as exc:
        note = f"Profile check: RULES ERROR — {os.path.basename(prof)} @ {stamp}: {exc}"
        if not quiet:
            print("  " + note)
        return {prof: [f"RULES ERROR: {exc}"]}, note
    bad, n = {}, 0
    for r in rows:
        folder = str(r.get("folder_path", ""))
        if not folder or not os.path.isdir(folder):
            continue
        for doc, fails, _warns in vp.check_folder(folder, ptext, rules):
            n += 1
            if fails:
                bad.setdefault(folder, []).extend(
                    f"{os.path.basename(doc)}: {f}" for f in fails)
    note = (f"Profile check: {n} docs, {sum(len(v) for v in bad.values())} failures, "
            f"{os.path.basename(prof)} @ {stamp}")
    if not quiet:
        state = f"FAILED for {len(bad)} role(s)" if bad else "OK"
        print(f"  profile check {state} ({n} docs against {os.path.basename(prof)})")
    return bad, note


def write_apply_today(rows, dest, date, blocked=None, note=""):
    """The list of what to send. Line 3 is the provenance note, ALWAYS: without it a day
    folder cannot distinguish "the gate passed" from "the gate never ran"."""
    out = [f"# Apply today — {date}", "", note or "Profile check: not run", ""]
    blocked = blocked or {}
    if blocked:
        out += ["## Fix before sending", "",
                f"{len(blocked)} role(s) have a document that contradicts the profile. They are "
                "still in their folders so you can correct them, but do not send them as they "
                "stand. Re-run `scripts/validate_profile.py` after fixing.", ""]
        for folder, fails in sorted(blocked.items()):
            out.append(f"- `{os.path.basename(folder)}`")
            out += [f"    - {f}" for f in fails]
        out.append("")
    out += [f"{len(rows)} application(s) listed." if rows else "_Nothing is ready for this date._",
            ""]
    for i, r in enumerate(rows, 1):
        closing = str(r.get("closing_date") or "").strip()
        out += [f"## {i}. {r.get('company', '')} — {r.get('role', '')}",
                f"- **Files:** `CV_{slug(r.get('company'))}_{slug(r.get('role'))}.docx` "
                f"(+ CoverLetter_…) in this folder",
                f"- **Pay:** {str(r.get('pay') or '').strip() or 'not stated'}"
                f"  ·  **Location:** {r.get('location', '') or 'not stated'}"
                + (f"  ·  **Closes:** {closing}" if closing else ""),
                f"- **Apply:** {r.get('link', '') or 'link missing — see notes.md'}",
                ""]
    path = os.path.join(dest, "APPLY-TODAY.md")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(out))
    return path


def write_xlsx(rows, dest, date):
    """One sheet of the day's roles — the thing you actually scan before applying."""
    wb = Workbook()
    ws = wb.active
    ws.title = date

    ws.append(SHEET_COLUMNS)
    for i, name in enumerate(SHEET_COLUMNS, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(i)].width = {
            "role": 40, "company": 30, "location": 22, "link": 46,
            "folder_path": 60, "pay": 22,
        }.get(name, 14)
    ws.freeze_panes = "A2"

    today = datetime.date.today()
    close_col = SHEET_COLUMNS.index("closing_date") + 1
    # Drafted first: those are the ones with a CV waiting to be sent.
    for r in sorted(rows, key=lambda x: (str(x.get("status")) != "Drafted",
                                         str(x.get("company", "")))):
        ws.append([safe_cell(r.get(c, "")) for c in SHEET_COLUMNS])
        cd = parse_date(r.get("closing_date"))
        if cd:
            days = (cd - today).days
            if days <= 7:
                ws.cell(row=ws.max_row, column=close_col).fill = URGENT_FILL
            elif days <= 14:
                ws.cell(row=ws.max_row, column=close_col).fill = SOON_FILL

    path = os.path.join(dest, f"{date}-roles.xlsx")
    try:
        wb.save(path)
    except (PermissionError, OSError) as e:
        sys.exit(f"Cannot write {path}: it is open in Excel or read-only ({e}). Close it and re-run.")
    return path


FINDINGS_SCAFFOLD = """# Daily bundle — {date}

**AI-assisted search — verify details with the employer.** Apply from this folder.

Everything for today is here: a tailored CV and cover letter per role, and `{date}-roles.xlsx`
listing them with pay, closing date, fit score and the apply link.

## READY TO APPLY

_(one section per role: lane, location, pay, closing date, apply link, why it fits, the files,
and anything to check before sending)_

## TRIAGED OUT

_(roles considered and dropped, each with the actual reason — so the same role is not re-found
and re-assessed next week)_

## Notes

_(what the run learned: connectors used, boards that were thin, anything to do differently)_
"""


def self_check():
    """One sample role in a temp workspace, bundled three times: clean, with a failing
    document, and with a broken rules block — line 3 must say which every time."""
    import subprocess
    import tempfile
    import cvgen
    import tracker
    assets = os.path.join(HERE, "..", "assets")
    ws = tempfile.mkdtemp()
    try:
        os.makedirs(os.path.join(ws, "profiles"))
        prof = os.path.join(ws, "profiles", "profile.md")
        shutil.copy(os.path.join(assets, "sample-profile.md"), prof)
        shutil.copy(os.path.join(assets, "sample-profile.blocks.md"),
                    cvgen.blocks_path_for(prof))
        ptext = io.open(prof, encoding="utf-8").read()
        rules = vp.parse_rules(ptext)
        apps = os.path.join(ws, "applications")
        folder = os.path.join(apps, "research", "2026-01-01_Acme_Analyst")
        os.makedirs(folder)
        cv = cvgen.build(ptext, rules, "research", "Analyst", "Summary.", ["- R."],
                         "A role.", cvgen.load_blocks(cvgen.blocks_path_for(prof)))
        with open(os.path.join(folder, "CV.md"), "w", encoding="utf-8", newline="\n") as f:
            f.write(cv)
        with open(os.path.join(folder, "CV.docx"), "wb") as f:
            f.write(b"PK")
        tracker.cmd_init(apps)
        tracker.cmd_add(apps, {"company": "Acme", "role": "Analyst", "folder_path": folder,
                               "category": "research", "logged_date": "2026-01-01"})
        dest = os.path.join(ws, "tasks", "daily", "2026-01-01")
        os.makedirs(dest)
        with open(os.path.join(dest, "company_rows.json"), "w") as f:
            f.write("{}")

        def run():
            subprocess.run([sys.executable, os.path.abspath(__file__), "--root", apps,
                            "--date", "2026-01-01", "--quiet"], check=True)
            with open(os.path.join(dest, "APPLY-TODAY.md"), encoding="utf-8") as f:
                return f.read().splitlines()[2]

        line3 = run()
        assert line3.startswith("Profile check: 1 docs, 0 failures, profile.md @ "), line3
        assert os.path.isfile(os.path.join(dest, "_work", "company_rows.json"))
        assert os.path.isfile(os.path.join(dest, "CV_Acme_Analyst.docx"))
        with open(os.path.join(folder, "CV.md"), "a", encoding="utf-8") as f:
            f.write("\nIELTS 7.0\n")
        assert "1 failures" in run()
        with open(prof, "w", encoding="utf-8", newline="\n") as f:
            f.write(ptext.replace("forbid: IELTS", "frobid: IELTS"))
        assert run().startswith("Profile check: RULES ERROR")
    finally:
        shutil.rmtree(ws, ignore_errors=True)
    print("daily_bundle self-check OK")
    return 0


def main():
    if "--self-check" in sys.argv:
        return self_check()
    ap = argparse.ArgumentParser(description="Assemble a dated apply-from-here bundle for one day.")
    ap.add_argument("--root", required=True, help="applications directory")
    ap.add_argument("--date", help="YYYY-MM-DD (default: today)")
    ap.add_argument("--tasks-dir", help="daily tasks dir (default: <workspace>/tasks/daily)")
    ap.add_argument("--status", default="Drafted,Skipped",
                    help="comma-separated statuses to include; 'all' for everything")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    apps = os.path.abspath(args.root)
    date = args.date or datetime.date.today().isoformat()
    if not parse_date(date):
        sys.exit(f"--date must be YYYY-MM-DD, got {date!r}")

    statuses = None
    if args.status and args.status.lower() != "all":
        statuses = {s.strip() for s in args.status.split(",") if s.strip()}

    tasks = args.tasks_dir or os.path.join(os.path.dirname(apps), "tasks", "daily")
    dest = os.path.join(os.path.abspath(tasks), date)
    os.makedirs(dest, exist_ok=True)

    rows = read_rows(apps, date, statuses)
    if not rows:
        print(f"No tracker rows logged on {date}"
              + (f" with status in {sorted(statuses)}" if statuses else "")
              + f". Created the empty bundle at {dest} anyway.")

    if not args.quiet:
        print(f"Bundling {len(rows)} role(s) for {date}")

    copied = copy_docs(rows, dest, args.quiet)
    blocked, note = profile_check(rows, apps, args.quiet)
    write_apply_today(rows, dest, date, blocked, note)
    xlsx = write_xlsx(rows, dest, date)
    tidy(dest, args.quiet)
    if not args.quiet:
        print(f"  sheet   {os.path.basename(xlsx)}")
        print("  wrote   APPLY-TODAY.md")

    findings = os.path.join(dest, "FINDINGS.md")
    if os.path.exists(findings):
        if not args.quiet:
            print("  kept    FINDINGS.md (already written — never overwritten)")
    else:
        with open(findings, "w", encoding="utf-8", newline="\n") as f:
            f.write(FINDINGS_SCAFFOLD.format(date=date))
        if not args.quiet:
            print("  wrote   FINDINGS.md (scaffold — fill it in)")

    drafted = sum(1 for r in rows if str(r.get("status")) == "Drafted")
    print(f"\n{len(rows)} role(s) · {drafted} ready to apply · {len(copied)} document(s)")
    print(dest)


if __name__ == "__main__":
    sys.exit(main())
