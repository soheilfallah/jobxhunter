#!/usr/bin/env python3
"""Application tracker — deterministic owner of tracker.xlsx + tracker.csv.

The skill must route ALL tracker state changes through this script so they are
deterministic, not model-improvised. Applied rows are filled green and locked
(sheet protection) so an applied record cannot be silently lost or altered.

Unique key for a row = folder_path (the per-job folder). Add is idempotent on it.

Usage:
  python tracker.py init   --root <applications_dir>
  python tracker.py add    --root <applications_dir> --data '<json>'
  python tracker.py update --root <applications_dir> --key <folder_path> --data '<json>'
  python tracker.py show   --root <applications_dir>

`--data` is a JSON object whose keys are any of COLUMNS. `add` fills logged_date
automatically if absent and defaults status to "Drafted". `update` sets fields on
the matching row; setting status to "Applied" stamps date_applied (if absent),
turns the row green, and locks it. Other status values stamp their matching date
column and recolour the row.
"""
import argparse
import csv
import datetime
import json
import os
import sys

# Preflight before importing openpyxl so a missing dep fails LOUDLY with the exact
# fix, instead of a bare sys.exit that let callers half-commit (folder but no row).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _lib import preflight, safe_cell  # noqa: E402
preflight([("openpyxl", "openpyxl", "tracker.xlsx read/write (tracker.py)")])

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import PatternFill, Font, Protection, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

COLUMNS = [
    "logged_date", "category", "company", "role", "location", "link", "source",
    "ats_platform", "pay", "level_used", "status", "date_applied",
    "date_interviewed", "date_rejected", "date_offer", "follow_up",
    "cv_path", "cover_letter_path", "folder_path", "notes",
    # appended (not inserted) so existing trackers keep their column positions:
    "closing_date", "fit_score",
]

# status -> the date column it stamps (if any)
STATUS_DATE = {
    "Applied": "date_applied",
    "Interview": "date_interviewed",
    "Interviewed": "date_interviewed",
    "Rejected": "date_rejected",
    "Offer": "date_offer",
}

STATUS_FILL = {
    "Applied":     "C6EFCE",  # green  — locked, final
    "Interview":   "BDD7EE",  # blue
    "Interviewed": "BDD7EE",
    "Offer":       "FFD966",  # gold
    "Rejected":    "F4CCCC",  # red
    "Skipped":     "D9D9D9",  # grey
    "Not applied": "D9D9D9",
    "Cold-emailed":"E4DFEC",  # light purple — speculative outreach sent
    "Replied":     "DDEBF7",  # pale blue — got a response to outreach
    "Drafted":     None,      # no fill
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def xlsx_path(root):
    return os.path.join(root, "tracker.xlsx")


def csv_path(root):
    return os.path.join(root, "tracker.csv")


def _today():
    # ISO date; caller may override via --data for reproducibility.
    return datetime.date.today().isoformat()


def _normkey(path):
    """Normalise a folder_path for comparison: unify separators, drop trailing
    slash, case-fold (Windows paths arrive with mixed / and \\)."""
    if path is None:
        return ""
    p = str(path).strip().replace("\\", "/").rstrip("/")
    return os.path.normcase(p)


def _new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "applications"
    ws.append(COLUMNS)
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")
        # sensible-ish widths
        width = {"notes": 40, "role": 26, "company": 22, "folder_path": 40,
                 "cv_path": 30, "cover_letter_path": 30, "link": 30}.get(name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    return wb, ws


def _load(root):
    p = xlsx_path(root)
    if not os.path.exists(p):
        sys.exit(f"No tracker at {p}. Run: tracker.py init --root {root}")
    wb = load_workbook(p)
    ws = wb["applications"] if "applications" in wb.sheetnames else wb.active
    return wb, ws


def _rows_as_dicts(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, len(COLUMNS) + 1)):
            continue
        d = {COLUMNS[c - 1]: (ws.cell(row=r, column=c).value or "") for c in range(1, len(COLUMNS) + 1)}
        d["_row"] = r
        rows.append(d)
    return rows


def _apply_row_style(ws, r, status):
    fill_hex = STATUS_FILL.get(status, None)
    locked = status in ("Applied",)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=c)
        if fill_hex:
            cell.fill = PatternFill("solid", fgColor=fill_hex)
        else:
            cell.fill = PatternFill(fill_type=None)
        # every data cell unlocked by default; only applied rows locked.
        cell.protection = Protection(locked=locked)


def _reprotect(ws):
    # Unlock header + everything, then rows re-locked individually in _apply_row_style.
    # Enable sheet protection so locked cells are actually enforced.
    ws.protection.sheet = True
    ws.protection.enable()
    # allow the user to still select/sort; they just can't edit locked (applied) cells
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def _write_csv(ws, root):
    with open(csv_path(root), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for d in _rows_as_dicts(ws):
            w.writerow([safe_cell(d[c]) for c in COLUMNS])


def _save(wb, ws, root):
    _reprotect(ws)
    wb.save(xlsx_path(root))
    _write_csv(ws, root)


def cmd_init(root):
    os.makedirs(root, exist_ok=True)
    if os.path.exists(xlsx_path(root)):
        print(f"Tracker already exists at {xlsx_path(root)} — leaving intact.")
        return
    wb, ws = _new_workbook()
    # header cells locked by default (fine); data rows added later control their own lock
    _save(wb, ws, root)
    print(f"Initialised tracker: {xlsx_path(root)} (+ csv mirror)")


def cmd_add(root, data):
    wb, ws = _load(root)
    data = dict(data)
    data.setdefault("logged_date", _today())
    data.setdefault("status", "Drafted")
    key = _normkey(data.get("folder_path", ""))
    if key:
        for d in _rows_as_dicts(ws):
            if _normkey(d["folder_path"]) == key:
                print(f"Row already exists for folder_path={data.get('folder_path')} (row {d['_row']}); use update.")
                return
    row = [safe_cell(data.get(col, "")) for col in COLUMNS]
    ws.append(row)
    r = ws.max_row
    _apply_row_style(ws, r, data.get("status", "Drafted"))
    _save(wb, ws, root)
    print(f"Added row {r}: {data.get('company','?')} / {data.get('role','?')} [{data.get('status')}]")


def cmd_update(root, key, data):
    wb, ws = _load(root)
    target = None
    nk = _normkey(key)
    for d in _rows_as_dicts(ws):
        if _normkey(d["folder_path"]) == nk:
            target = d
            break
    if not target:
        sys.exit(f"No row with folder_path={key}")
    r = target["_row"]
    # guard: applied rows are final. Refuse silent overwrite unless --force via data.
    if str(target.get("status")) == "Applied" and data.get("status") not in (None, "Applied"):
        if not data.get("_force"):
            sys.exit(f"Row {r} is Applied (locked/final). Pass \"_force\": true in --data to override.")
    data = {k: v for k, v in data.items() if k != "_force"}
    for k, v in data.items():
        if k in COLUMNS:
            ws.cell(row=r, column=COLUMNS.index(k) + 1, value=safe_cell(v))
    new_status = data.get("status", target.get("status"))
    # auto-stamp the matching date column if a status implies one and it's empty
    date_col = STATUS_DATE.get(str(new_status))
    if date_col and not ws.cell(row=r, column=COLUMNS.index(date_col) + 1).value:
        ws.cell(row=r, column=COLUMNS.index(date_col) + 1, value=safe_cell(data.get(date_col, _today())))
    _apply_row_style(ws, r, str(new_status))
    _save(wb, ws, root)
    print(f"Updated row {r} -> status={new_status}")


def cmd_show(root):
    _, ws = _load(root)
    rows = _rows_as_dicts(ws)
    print(f"{len(rows)} row(s) in {xlsx_path(root)}")
    for d in rows:
        print(f"  [{d['status']:<11}] {d['category']:<20} {d['company']:<22} {d['role']}")


# status precedence for choosing which duplicate to keep (higher = keep)
STATUS_RANK = {"Offer": 6, "Interviewed": 5, "Interview": 5, "Rejected": 4,
               "Applied": 4, "Replied": 3, "Cold-emailed": 2, "Drafted": 1,
               "Skipped": 0}
# post-application states are NEVER auto-deleted as a duplicate
PROTECTED = {"Applied", "Interview", "Interviewed", "Offer", "Rejected"}


def _dedupe_key(d):
    """Canonical job key: normalised link, else folder_path. Never the folder slug alone."""
    from build_seen_ledger import canonical_key
    return canonical_key(d.get("link", ""), d.get("folder_path", ""))


def cmd_dedupe(root, apply=False):
    wb, ws = _load(root)
    rows = _rows_as_dicts(ws)
    groups = {}
    for d in rows:
        k = _dedupe_key(d)
        if k:
            groups.setdefault(k, []).append(d)

    to_delete = set()      # _row numbers to drop
    merges = {}            # _row -> {col: value} to backfill into an unprotected keeper
    warnings = []
    for k, grp in groups.items():
        if len(grp) < 2:
            continue
        protected = [d for d in grp if str(d.get("status")) in PROTECTED]
        rank = lambda d: (STATUS_RANK.get(str(d.get("status")), 0), -d["_row"])
        if protected:
            keeper = max(protected, key=rank)
            if len(protected) > 1:
                warnings.append(f"  ! {k}: {len(protected)} post-application rows "
                                f"(rows {[d['_row'] for d in protected]}) — NOT auto-merged, review by hand")
            # only delete UNprotected duplicates; never touch a protected/Applied row
            for d in grp:
                if d is not keeper and str(d.get("status")) not in PROTECTED:
                    to_delete.add(d["_row"])
        else:
            keeper = max(grp, key=rank)
            backfill = {}
            for d in grp:
                if d is keeper:
                    continue
                to_delete.add(d["_row"])
                for col in COLUMNS:  # fill keeper's blanks from the dup
                    if not keeper.get(col) and d.get(col):
                        backfill.setdefault(col, d[col])
            if backfill:
                merges[keeper["_row"]] = backfill

    dup_count = len(to_delete)
    print(f"Dedupe scan: {len(rows)} rows, {len(groups)} unique keys, "
          f"{dup_count} duplicate row(s) to remove.")
    for w in warnings:
        print(w)
    if dup_count == 0 and not merges:
        print("Nothing to dedupe.")
        return
    for r in sorted(to_delete):
        d = next(x for x in rows if x["_row"] == r)
        print(f"  - drop row {r}: [{d.get('status')}] {d.get('company')} / {d.get('role')}")

    if not apply:
        print("\nDRY RUN — re-run with --apply to rewrite the tracker.")
        return

    # Rebuild the sheet from survivors (keeps ordering, reapplies styles/locks).
    survivors = [d for d in rows if d["_row"] not in to_delete]
    for d in survivors:
        if d["_row"] in merges:
            d.update(merges[d["_row"]])
    new_wb, new_ws = _new_workbook()
    for d in survivors:
        new_ws.append([safe_cell(d.get(c, "")) for c in COLUMNS])
        _apply_row_style(new_ws, new_ws.max_row, str(d.get("status", "Drafted")))
    _save(new_wb, new_ws, root)
    print(f"\nApplied: removed {dup_count} duplicate row(s); {len(survivors)} rows remain.")


# priority buckets for the day-to-day worklist (lower sorts first)
PRIORITY_BUCKET = {
    "Drafted": 0,                                            # actionable — tailor/apply
    "Cold-emailed": 1, "Replied": 1, "Interview": 1,
    "Interviewed": 1, "Offer": 1,                            # in-progress
    "Applied": 2,                                            # done for now — sinks
    "Rejected": 3, "Skipped": 3, "Not applied": 3,           # dead / watch — bottom
}
PRIORITY_VIEW_COLUMNS = ["status", "company", "role", "location",
                         "closing_date", "fit_score", "link"]


def _parse_date(s):
    try:
        return datetime.date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def _fit_num(v):
    """Fit/recruiter score -> float for sorting (higher = better). Blank -> -1."""
    try:
        return float(str(v).split("/")[0].strip())
    except (ValueError, AttributeError):
        return -1.0


def cmd_priority_view(root):
    _, ws = _load(root)
    rows = _rows_as_dicts(ws)
    today = datetime.date.today()
    far = datetime.date(9999, 12, 31)

    def sort_key(d):
        bucket = PRIORITY_BUCKET.get(str(d.get("status")), 2)
        close = _parse_date(d.get("closing_date")) or far      # no date -> last
        fit = -_fit_num(d.get("fit_score"))                    # higher fit first
        return (bucket, close, fit)

    rows.sort(key=sort_key)

    wb = Workbook()
    out = wb.active
    out.title = "worklist"
    out.append(PRIORITY_VIEW_COLUMNS)
    for i, name in enumerate(PRIORITY_VIEW_COLUMNS, start=1):
        c = out.cell(row=1, column=i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        width = {"role": 30, "company": 24, "location": 20, "link": 40}.get(name, 14)
        out.column_dimensions[get_column_letter(i)].width = width
    out.freeze_panes = "A2"

    red = PatternFill("solid", fgColor="F4CCCC")
    amber = PatternFill("solid", fgColor="FCE4D6")
    close_col = PRIORITY_VIEW_COLUMNS.index("closing_date") + 1
    for d in rows:
        out.append([safe_cell(d.get(c, "")) for c in PRIORITY_VIEW_COLUMNS])
        r = out.max_row
        cd = _parse_date(d.get("closing_date"))
        if cd:
            days = (cd - today).days
            if days <= 7:
                out.cell(row=r, column=close_col).fill = red      # closes within a week
            elif days <= 14:
                out.cell(row=r, column=close_col).fill = amber

    path = os.path.join(root, "tracker-priority.xlsx")
    wb.save(path)
    print(f"Prioritised worklist: {path} ({len(rows)} rows; Drafted first, "
          f"soonest-closing next, closing≤7d in red). Full tracker stays the system of record.")


def main():
    ap = argparse.ArgumentParser(description="Application tracker (xlsx + csv).")
    sub = ap.add_subparsers(dest="cmd", required=True)
    for name in ("init", "add", "update", "show", "dedupe", "priority-view"):
        p = sub.add_parser(name)
        p.add_argument("--root", required=True, help="applications directory")
        if name in ("add", "update"):
            p.add_argument("--data", required=True, help="JSON object of column values")
        if name == "update":
            p.add_argument("--key", required=True, help="folder_path of the row to update")
        if name == "dedupe":
            p.add_argument("--apply", action="store_true", help="rewrite (default: dry run)")
    args = ap.parse_args()
    data = json.loads(args.data) if getattr(args, "data", None) else {}
    if args.cmd == "init":
        cmd_init(args.root)
    elif args.cmd == "add":
        cmd_add(args.root, data)
    elif args.cmd == "update":
        cmd_update(args.root, args.key, data)
    elif args.cmd == "show":
        cmd_show(args.root)
    elif args.cmd == "dedupe":
        cmd_dedupe(args.root, apply=args.apply)
    elif args.cmd == "priority-view":
        cmd_priority_view(args.root)


if __name__ == "__main__":
    main()
